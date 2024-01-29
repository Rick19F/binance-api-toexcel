import openpyxl
from openpyxl.chart import LineChart, Reference
from datetime import datetime
import os.path
import requests

# Binance API endpoint for getting current ticker price
BINANCE_API_URL = 'https://api.binance.com/api/v3/ticker/price'
# Symbols for Bitcoin and Ethereum
BTC_SYMBOL = 'BTCUSDT'
ETH_SYMBOL = 'ETHUSDT'
# File to write results
OUTPUT_FILE = 'crypto_prices.xlsx'
# Name of the sheet in Excel
SHEET_NAME = 'Crypto Prices'
# Chart locations
BTC_CHART_CELL = 'D2'
ETH_CHART_CELL = 'D20'

def get_crypto_price(symbol):
    response = requests.get(BINANCE_API_URL, params={'symbol': symbol})
    data = response.json()
    crypto_price = float(data['price'])
    return crypto_price

def write_to_excel(btc_price, eth_price):
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    sheet = wb[SHEET_NAME]
    sheet.append([datetime.now(), btc_price, eth_price])
    
    # Apply date format to the first column (dates)
    sheet.column_dimensions['A'].width = 20
    for row in sheet.iter_rows(min_row=sheet.min_row, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = 'YYYY-MM-DD HH:MM:SS'
    
    # Apply currency format to the second and third columns (BTC and ETH prices)
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 15
    for row in sheet.iter_rows(min_row=sheet.min_row, max_row=sheet.max_row, min_col=2, max_col=3):
        for cell in row:
            cell.number_format = '$#,##0.00'
    
    wb.save(OUTPUT_FILE)

def update_charts():
    if os.path.isfile(OUTPUT_FILE):
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        if SHEET_NAME in wb.sheetnames:
            sheet = wb[SHEET_NAME]

            btc_chart = LineChart()
            btc_chart.title = "Bitcoin Prices"
            btc_chart.y_axis.title = "Price (USDT)"
            btc_chart.x_axis.title = "Time"
            btc_chart.marker = "circle"

            eth_chart = LineChart()
            eth_chart.title = "Ethereum Prices"
            eth_chart.y_axis.title = "Price (USDT)"
            eth_chart.x_axis.title = "Time"
            eth_chart.marker = "circle"

            data = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
            dates = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
            btc_data = Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=sheet.max_row)
            eth_data = Reference(sheet, min_col=3, min_row=1, max_col=3, max_row=sheet.max_row)

            btc_chart.add_data(btc_data, titles_from_data=True)
            btc_chart.set_categories(dates)

            eth_chart.add_data(eth_data, titles_from_data=True)
            eth_chart.set_categories(dates)

            sheet.add_chart(btc_chart, BTC_CHART_CELL)
            sheet.add_chart(eth_chart, ETH_CHART_CELL)

        wb.save(OUTPUT_FILE)

def main():
    btc_price = get_crypto_price(BTC_SYMBOL)
    eth_price = get_crypto_price(ETH_SYMBOL)
    print("Current Bitcoin Price (in USDT):", btc_price)
    print("Current Ethereum Price (in USDT):", eth_price)
    write_to_excel(btc_price, eth_price)
    update_charts()

if __name__ == "__main__":
    main()
